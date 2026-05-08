"""
ES5坐标一致性验证工具
从Excel文件读取采集数据，验证每行Y坐标、每列X坐标是否一致
用法: python validate_coordinates.py <excel文件路径> [--tolerance 0.01]
"""

import sys
import argparse
from openpyxl import load_workbook


def validate_coordinates_from_file(filepath, tolerance=0.01):
    """从Excel文件读取数据并验证坐标一致性"""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"错误: 无法打开文件 {filepath}: {e}")
        return False

    ws = wb.active

    # 读取表头确认格式
    headers = [cell.value for cell in ws[1]]
    expected = ['序号', '行号', '列号', '屏幕X', '屏幕Y', '坐标X', '坐标Y', '坐标Z', '声压级(dB)']
    if headers != expected:
        print(f"警告: 表头不匹配，期望 {expected}，实际 {headers}")
        print("将尝试按列位置读取（F=坐标X, G=坐标Y, B=行号, C=列号, A=序号）")

    # 读取数据
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        data.append({
            'index': row[0],
            'row': row[1],
            'col': row[2],
            'coord_x': row[5] if len(row) > 5 else 0.0,
            'coord_y': row[6] if len(row) > 6 else 0.0,
        })
    wb.close()

    if not data:
        print("错误: 文件中没有数据！")
        return False

    print(f"已读取 {len(data)} 个数据点\n")

    # 推断行列数
    rows = max(d['row'] for d in data)
    cols = max(d['col'] for d in data)

    has_error = False

    print("=" * 60)
    print("坐标一致性验证")
    print("=" * 60)

    # 按行检查Y轴坐标
    print(f"\n[1] 按行检查Y轴坐标一致性 (容差: {tolerance}):")
    print("-" * 60)

    for row_num in range(1, rows + 1):
        row_points = [d for d in data if d['row'] == row_num]
        if not row_points:
            continue

        ref_y = row_points[0]['coord_y']
        inconsistent = [p for p in row_points if abs(p['coord_y'] - ref_y) > tolerance]

        if inconsistent:
            has_error = True
            print(f"\n  ⚠ 行 {row_num} Y轴坐标不一致 (参考值: {ref_y:.2f}):")
            for pt in inconsistent:
                print(f"    - 序号:{pt['index']}, 行:{pt['row']}, 列:{pt['col']}, "
                      f"Y={pt['coord_y']:.2f} (偏差: {abs(pt['coord_y'] - ref_y):.2f})")
        else:
            print(f"  ✓ 行 {row_num}: Y={ref_y:.2f} (一致)")

    # 按列检查X轴坐标
    print(f"\n[2] 按列检查X轴坐标一致性 (容差: {tolerance}):")
    print("-" * 60)

    for col_num in range(1, cols + 1):
        col_points = [d for d in data if d['col'] == col_num]
        if not col_points:
            continue

        ref_x = col_points[0]['coord_x']
        inconsistent = [p for p in col_points if abs(p['coord_x'] - ref_x) > tolerance]

        if inconsistent:
            has_error = True
            print(f"\n  ⚠ 列 {col_num} X轴坐标不一致 (参考值: {ref_x:.2f}):")
            for pt in inconsistent:
                print(f"    - 序号:{pt['index']}, 行:{pt['row']}, 列:{pt['col']}, "
                      f"X={pt['coord_x']:.2f} (偏差: {abs(pt['coord_x'] - ref_x):.2f})")
        else:
            print(f"  ✓ 列 {col_num}: X={ref_x:.2f} (一致)")

    # 总结
    print("\n" + "=" * 60)
    if has_error:
        print("验证结果: ⚠ 发现坐标不一致问题，请检查上方详细信息")
    else:
        print("验证结果: ✓ 所有坐标一致性检查通过")
    print("=" * 60 + "\n")

    return not has_error


def main():
    parser = argparse.ArgumentParser(description='ES5坐标一致性验证工具')
    parser.add_argument('filepath', help='Excel文件路径')
    parser.add_argument('--tolerance', type=float, default=0.01, help='坐标容差 (默认: 0.01)')
    args = parser.parse_args()

    result = validate_coordinates_from_file(args.filepath, args.tolerance)
    sys.exit(0 if result else 1)


if __name__ == '__main__':
    # main()
    validate_coordinates_from_file(r"E:\AudioSimulationSnap\ES5数据_20260508_110508.xlsx")