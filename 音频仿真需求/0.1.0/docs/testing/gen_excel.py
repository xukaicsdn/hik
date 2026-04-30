import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式定义 =====
header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
title_font = Font(name='微软雅黑', bold=True, size=14, color='1F4E79')
normal_font = Font(name='微软雅黑', size=10)
p0_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
p1_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
p2_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

def style_row(ws, row, cols, priority_col=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.alignment = left_align if c > 2 else center_align
        cell.border = thin_border
    if priority_col:
        pval = ws.cell(row=row, column=priority_col).value
        if pval == 'P0':
            ws.cell(row=row, column=priority_col).fill = p0_fill
        elif pval == 'P1':
            ws.cell(row=row, column=priority_col).fill = p1_fill
        elif pval == 'P2':
            ws.cell(row=row, column=priority_col).fill = p2_fill

# ==============================
# Sheet 1: 功能测试用例
# ==============================
ws1 = wb.active
ws1.title = '功能测试用例'

ws1.merge_cells('A1:H1')
ws1.cell(row=1, column=1, value='音频仿真工具 v0.1.0 功能测试用例').font = title_font
ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 35

headers = ['用例编号', '测试场景', '测试步骤', '输入数据', '预期结果', '优先级', '测试结果', '备注']
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, len(headers))

test_cases = [
    ('TC-001', '3D窗口场景操作', '1. 打开3D窗口\n2. 按住鼠标左键拖动', '鼠标左键拖动', '场景顺时针/逆时针旋转', 'P0', '', ''),
    ('TC-002', '3D窗口场景操作', '1. 打开3D窗口\n2. 滚动鼠标滚轮', '滚轮滚动', '场景放大/缩小', 'P0', '', ''),
    ('TC-003', '3D窗口场景操作', '1. 打开3D窗口\n2. 按住鼠标中键拖动', '鼠标中键拖动', '场景平移', 'P0', '', ''),
    ('TC-004', '3D窗口场景操作', '1. 打开3D窗口\n2. 双击场景对象', '鼠标双击', '聚焦到选中对象', 'P1', '', ''),
    ('TC-005', '3D窗口场景操作', '1. 打开3D窗口\n2. 右键拖动', '鼠标右键拖动', '场景旋转', 'P1', '', ''),
    ('TC-006', '图纸导入功能', '1. 点击菜单"文件-导入图纸"\n2. 选择DXF文件\n3. 点击打开', '有效DXF文件', '成功导入图纸，场景显示导入的图形', 'P0', '', ''),
    ('TC-007', '图纸导入功能', '1. 点击菜单"文件-导入图纸"\n2. 选择DWG文件\n3. 点击打开', '有效DWG文件', '成功导入图纸，场景显示导入的图形', 'P0', '', ''),
    ('TC-008', '图纸导入功能', '1. 点击菜单"文件-导入图纸"\n2. 选择无效文件\n3. 点击打开', '损坏的DXF文件', '显示错误提示，导入失败', 'P1', '', ''),
    ('TC-009', '图纸导入功能', '1. 点击菜单"文件-导入图纸"\n2. 选择非CAD文件\n3. 点击打开', 'TXT文件', '显示格式不支持提示', 'P1', '', ''),
    ('TC-010', '图纸导入功能', '1. 导入图纸后\n2. 点击"撤销"', '无', '图纸被撤销，场景恢复原状', 'P2', '', ''),
    ('TC-011', '图纸导入功能', '1. 导入图纸后\n2. 点击"重做"', '无', '图纸被恢复', 'P2', '', ''),
    ('TC-012', '声源管理', '1. 点击工具栏"添加声源"\n2. 在场景中点击放置位置', '鼠标点击位置', '在指定位置创建声源图标', 'P0', '', ''),
    ('TC-013', '声源管理', '1. 选中已有声源\n2. 修改属性面板参数', '声压级: 80dB\n频率: 1000Hz', '属性更新，声源参数改变', 'P0', '', ''),
    ('TC-014', '声源管理', '1. 选中已有声源\n2. 按 Delete 键', '无', '声源被删除', 'P0', '', ''),
    ('TC-015', '声源管理', '1. 创建多个声源\n2. 选中其中一个', '无', '选中声源高亮显示，其他不变', 'P0', '', ''),
    ('TC-016', '声源管理', '1. 选中声源\n2. 拖动到新位置', '鼠标拖动', '声源位置更新', 'P0', '', ''),
    ('TC-017', '声源管理', '1. 创建声源\n2. 修改声源类型', '类型: 点声源/线声源/面声源', '声源类型切换，图标样式改变', 'P1', '', ''),
    ('TC-018', '声源管理', '1. 创建声源\n2. 设置指向性参数', '指向性角度: 45度', '指向性可视化显示正确', 'P1', '', ''),
    ('TC-019', '表面材质配置', '1. 选中场景中的表面对象\n2. 打开材质属性面板', '无', '显示表面材质属性', 'P0', '', ''),
    ('TC-020', '表面材质配置', '1. 选中表面\n2. 选择材质类型', '材质: 混凝土/金属/玻璃/木材', '表面视觉材质改变', 'P0', '', ''),
    ('TC-021', '表面材质配置', '1. 选中表面\n2. 设置吸声系数', '吸声系数: 0.5', '吸声系数参数保存成功', 'P0', '', ''),
    ('TC-022', '表面材质配置', '1. 选中表面\n2. 设置反射系数', '反射系数: 0.3', '反射系数参数保存成功', 'P0', '', ''),
    ('TC-023', '表面材质配置', '1. 选中表面\n2. 设置透射系数', '透射系数: 0.1', '透射系数参数保存成功', 'P0', '', ''),
    ('TC-024', '表面材质配置', '1. 选中表面\n2. 设置粗糙度', '粗糙度: 0.02', '粗糙度参数保存成功', 'P1', '', ''),
    ('TC-025', '表面材质配置', '1. 选中表面\n2. 设置密度', '密度: 2500 kg/m3', '密度参数保存成功', 'P1', '', ''),
    ('TC-026', '听音面设置', '1. 点击"添加听音面"\n2. 在场景中绘制区域', '鼠标绘制', '创建听音面区域', 'P0', '', ''),
    ('TC-027', '听音面设置', '1. 选中听音面\n2. 设置高度', '高度: 1.2m', '听音面高度更新', 'P0', '', ''),
    ('TC-028', '听音面设置', '1. 选中听音面\n2. 设置网格密度', '网格: 0.5m', '网格密度更新', 'P1', '', ''),
    ('TC-029', '听音面设置', '1. 选中听音面\n2. 调整位置和大小', '拖动操作', '听音面几何参数更新', 'P0', '', ''),
    ('TC-030', '听音面设置', '1. 选中听音面\n2. 按 Delete 键', '无', '听音面被删除', 'P0', '', ''),
    ('TC-031', '音频仿真计算', '1. 配置声源和表面\n2. 点击"开始仿真"', '无', '仿真开始，进度条显示', 'P0', '', ''),
    ('TC-032', '音频仿真计算', '1. 仿真过程中\n2. 点击"取消"', '无', '仿真取消，进度停止', 'P1', '', ''),
    ('TC-033', '音频仿真计算', '1. 仿真完成后\n2. 查看结果', '无', '显示仿真结果数据', 'P0', '', ''),
    ('TC-034', '音频仿真计算', '1. 不设置声源\n2. 点击仿真', '无', '提示至少需要一个声源', 'P1', '', ''),
    ('TC-035', '音频仿真计算', '1. 不设置听音面\n2. 点击仿真', '无', '提示至少需要一个听音面', 'P1', '', ''),
    ('TC-036', '音频仿真计算', '1. 仿真完成后\n2. 导出结果', '无', '生成仿真报告文件', 'P1', '', ''),
    ('TC-037', '仿真结果展示', '1. 仿真完成后\n2. 查看声压级分布图', '无', '显示彩色云图', 'P0', '', ''),
    ('TC-038', '仿真结果展示', '1. 仿真完成后\n2. 查看频谱分析', '无', '显示频率响应曲线', 'P1', '', ''),
    ('TC-039', '仿真结果展示', '1. 仿真完成后\n2. 查看混响时间', '无', '显示RT60数据', 'P1', '', ''),
    ('TC-040', '仿真结果展示', '1. 查看结果时\n2. 切换颜色映射', '颜色: 冷色/暖色/灰度', '颜色方案切换', 'P2', '', ''),
    ('TC-041', '仿真结果展示', '1. 查看结果时\n2. 调整显示阈值', '阈值: 40-80dB', '只显示阈值范围内的数据', 'P2', '', ''),
    ('TC-042', '界面基础功能', '1. 点击菜单"文件-新建"', '无', '创建新项目，场景清空', 'P0', '', ''),
    ('TC-043', '界面基础功能', '1. 点击菜单"文件-打开"', '有效项目文件', '加载项目，场景恢复', 'P0', '', ''),
    ('TC-044', '界面基础功能', '1. 点击菜单"文件-保存"', '无', '保存当前项目', 'P0', '', ''),
    ('TC-045', '界面基础功能', '1. 点击菜单"文件-另存为"', '新文件路径', '项目保存到新位置', 'P1', '', ''),
    ('TC-046', '界面基础功能', '1. 点击菜单"编辑-撤销"', '无', '撤销上一步操作', 'P1', '', ''),
    ('TC-047', '界面基础功能', '1. 点击菜单"编辑-重做"', '无', '重做上一步操作', 'P1', '', ''),
    ('TC-048', '界面基础功能', '1. 点击菜单"视图-重置视角"', '无', '场景视角恢复默认', 'P1', '', ''),
    ('TC-049', '界面基础功能', '1. 点击菜单"视图-全屏显示"', '无', '切换到全屏模式', 'P2', '', ''),
    ('TC-050', '界面基础功能', '1. 点击菜单"帮助-关于"', '无', '显示版本信息', 'P2', '', ''),
]

row = 4
for tc in test_cases:
    for i, val in enumerate(tc, 1):
        ws1.cell(row=row, column=i, value=val)
    style_row(ws1, row, len(headers), priority_col=6)
    ws1.row_dimensions[row].height = 45
    row += 1

col_widths = [10, 16, 40, 22, 30, 8, 10, 12]
for i, w in enumerate(col_widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ==============================
# Sheet 2: 性能测试
# ==============================
ws2 = wb.create_sheet('性能测试')
ws2.merge_cells('A1:E1')
ws2.cell(row=1, column=1, value='音频仿真工具 v0.1.0 性能测试').font = title_font
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

perf_headers = ['编号', '性能指标', '目标值', '测试方法', '测试结果']
for i, h in enumerate(perf_headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, len(perf_headers))

perf_data = [
    ('PT-001', '软件启动时间', '<= 5 秒', '秒表计时，从点击到界面完全加载', ''),
    ('PT-002', '3D场景旋转帧率', '>= 30 FPS', '使用帧率监测工具', ''),
    ('PT-003', '图纸导入时间（10MB文件）', '<= 10 秒', '秒表计时', ''),
    ('PT-004', '仿真计算界面响应', '界面不卡顿', '主线程不被阻塞', ''),
    ('PT-005', '内存占用（空闲状态）', '<= 500 MB', '任务管理器监测', ''),
    ('PT-006', '内存占用（仿真中）', '<= 2 GB', '任务管理器监测', ''),
    ('PT-007', '大图纸加载', '无崩溃/无内存溢出', '加载 50MB+ CAD 文件', ''),
]

row = 4
for d in perf_data:
    for i, val in enumerate(d, 1):
        ws2.cell(row=row, column=i, value=val)
    style_row(ws2, row, len(perf_headers))
    ws2.row_dimensions[row].height = 30
    row += 1

perf_widths = [10, 28, 22, 35, 12]
for i, w in enumerate(perf_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# ==============================
# Sheet 3: 边界测试
# ==============================
ws3 = wb.create_sheet('边界测试')
ws3.merge_cells('A1:E1')
ws3.cell(row=1, column=1, value='音频仿真工具 v0.1.0 边界测试').font = title_font
ws3.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

bound_headers = ['编号', '边界条件', '测试输入', '预期行为', '测试结果']
for i, h in enumerate(bound_headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(bound_headers))

bound_data = [
    ('BT-001', '声源数量上限', '创建 100+ 个声源', '提示数量限制或性能下降警告', ''),
    ('BT-002', '表面数量上限', '创建 1000+ 个表面', '正常处理或提示限制', ''),
    ('BT-003', '吸声系数下限', '设置为 0.0', '计算正常处理', ''),
    ('BT-004', '吸声系数上限', '设置为 1.0', '计算正常处理', ''),
    ('BT-005', '吸声系数越界', '设置为 1.5', '提示输入无效', ''),
    ('BT-006', '频率参数下限', '设置为 20 Hz', '正常处理（人耳最低频率）', ''),
    ('BT-007', '频率参数上限', '设置为 20000 Hz', '正常处理（人耳最高频率）', ''),
    ('BT-008', '频率参数越界', '设置为 -100 Hz', '提示输入无效', ''),
    ('BT-009', '坐标输入边界', '输入极大值坐标', '正常处理或提示越界', ''),
    ('BT-010', '空项目保存', '不添加任何对象直接保存', '允许保存空项目', ''),
    ('BT-011', '仿真区域边界', '听音面超出场景边界', '裁剪处理或提示警告', ''),
]

row = 4
for d in bound_data:
    for i, val in enumerate(d, 1):
        ws3.cell(row=row, column=i, value=val)
    style_row(ws3, row, len(bound_headers))
    ws3.row_dimensions[row].height = 30
    row += 1

bound_widths = [10, 20, 25, 28, 12]
for i, w in enumerate(bound_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ==============================
# Sheet 4: 异常测试
# ==============================
ws4 = wb.create_sheet('异常测试')
ws4.merge_cells('A1:E1')
ws4.cell(row=1, column=1, value='音频仿真工具 v0.1.0 异常测试').font = title_font
ws4.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 35

exc_headers = ['编号', '异常场景', '触发条件', '预期处理', '测试结果']
for i, h in enumerate(exc_headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, len(exc_headers))

exc_data = [
    ('ET-001', '文件损坏', '打开损坏的项目文件', '显示错误提示，不崩溃', ''),
    ('ET-002', '内存不足', '加载超大型 CAD 文件', '提示内存不足，不崩溃', ''),
    ('ET-003', '磁盘空间不足', '磁盘空间 < 100MB 时保存', '提示磁盘空间不足', ''),
    ('ET-004', '仿真参数异常', '声源和听音面距离为 0', '提示参数异常', ''),
    ('ET-005', '并发操作冲突', '快速连续点击多个按钮', '正确响应，无异常', ''),
    ('ET-006', '窗口最小化时仿真', '最小化窗口后开始仿真', '仿真继续运行', ''),
    ('ET-007', '显示器分辨率变更', '仿真过程中切换分辨率', '界面自适应调整', ''),
    ('ET-008', '显卡驱动异常', '模拟 OpenGL 错误', '显示错误提示，不崩溃', ''),
]

row = 4
for d in exc_data:
    for i, val in enumerate(d, 1):
        ws4.cell(row=row, column=i, value=val)
    style_row(ws4, row, len(exc_headers))
    ws4.row_dimensions[row].height = 30
    row += 1

exc_widths = [10, 22, 28, 25, 12]
for i, w in enumerate(exc_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w

# ==============================
# Sheet 5: 兼容性测试
# ==============================
ws5 = wb.create_sheet('兼容性测试')
ws5.merge_cells('A1:E1')
ws5.cell(row=1, column=1, value='音频仿真工具 v0.1.0 兼容性测试').font = title_font
ws5.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 35

comp_headers = ['编号', '测试项', '测试内容', '预期结果', '测试结果']
for i, h in enumerate(comp_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, len(comp_headers))

comp_data = [
    ('CT-001', '操作系统兼容性', 'Windows 10 专业版', '正常运行', ''),
    ('CT-002', '操作系统兼容性', 'Windows 10 家庭版', '正常运行', ''),
    ('CT-003', '操作系统兼容性', 'Windows 11 专业版', '正常运行', ''),
    ('CT-004', '显卡兼容性', 'NVIDIA 显卡', '3D 渲染正常', ''),
    ('CT-005', '显卡兼容性', 'AMD 显卡', '3D 渲染正常', ''),
    ('CT-006', '显卡兼容性', 'Intel 集成显卡', '3D 渲染正常（性能可能下降）', ''),
    ('CT-007', 'CAD 文件兼容性', 'AutoCAD 2018 生成的 DXF', '正常导入', ''),
    ('CT-008', 'CAD 文件兼容性', 'AutoCAD 2020 生成的 DWG', '正常导入', ''),
    ('CT-009', 'CAD 文件兼容性', '中文字符文件名', '正常处理', ''),
    ('CT-010', '多显示器', '连接多个显示器', '界面正常显示', ''),
]

row = 4
for d in comp_data:
    for i, val in enumerate(d, 1):
        ws5.cell(row=row, column=i, value=val)
    style_row(ws5, row, len(comp_headers))
    ws5.row_dimensions[row].height = 30
    row += 1

comp_widths = [10, 18, 28, 28, 12]
for i, w in enumerate(comp_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# ===== 保存 =====
output_path = r'E:\AudioSimulationSnap\音频仿真需求\0.1.0\docs\testing\test_plan_20260430.xlsx'
wb.save(output_path)
print(f'Excel saved to: {output_path}')
