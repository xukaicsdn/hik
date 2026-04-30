"""
    以EASE5为基准去匹配的，所以注意数据，如EASE有第8行3列，自研的没有尽量删掉，
    不然坐标匹配的不准
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Reference, ScatterChart, Series

# =========================
# 用户可直接在这里改参数（代码内指定）
# =========================
SELF_FILE_CONFIG = r"E:\AudioSimulationSnap\data\求解结果对比-直达声压（RoomMapping）\中型教室1\双声源\【中型教室1】【双声源，三分之一倍频程，A加权，直达声压】【1000HZ】- 声压数据_20260422_115551.xlsx"
EASE5_FILE_CONFIG = r"E:\AudioSimulationSnap\data\求解结果对比-直达声压（RoomMapping）\中型教室1\双声源\【中型教室1】【双声源，三分之一倍频程，A加权，直达声压】【1000HZ】- ES5数据_20260422_120334.xlsx"

# SELF_FILE_CONFIG = r"E:\AudioSimulationSnap\声压数据_20260430_145415.xlsx"
# EASE5_FILE_CONFIG = r"E:\AudioSimulationSnap\ES5数据_20260430_150223.xlsx"

SELF_SHEET_CONFIG = "声压数据"
EASE5_SHEET_CONFIG = "ES5数据"
OUT_DIR_CONFIG = "outputs/spl_compare"
MATCH_MODE_CONFIG = "grid"  # grid / repeat / unique
ANCHOR_CONFIG = "smaller"  # self / ease5 / smaller
ROW_COUNT_CONFIG = None
COL_COUNT_CONFIG = 20
Y_TOL_CONFIG = 1e-3
GENERATE_CHARTS_CONFIG = False


def _pick_col(columns: List[str], aliases: List[str]) -> str:
    lowered = {c.lower(): c for c in columns}
    for alias in aliases:
        if alias.lower() in lowered:
            return lowered[alias.lower()]

    for c in columns:
        c_low = c.lower()
        if any(alias.lower() in c_low for alias in aliases):
            return c
    raise KeyError(f"无法从列中匹配别名 {aliases}，当前列={columns}")


def detect_columns(df: pd.DataFrame) -> Dict[str, str]:
    cols = [str(c).strip() for c in df.columns]
    rename_map = {old: new for old, new in zip(df.columns, cols)}
    df.rename(columns=rename_map, inplace=True)

    x_col = _pick_col(cols, ["x", "x(m)", "x坐标", "坐标x"])
    y_col = _pick_col(cols, ["y", "y(m)", "y坐标", "坐标y"])
    z_col = _pick_col(cols, ["z", "z(m)", "z坐标", "坐标z"])
    spl_col = _pick_col(
        cols,
        [
            "spl",
            "lp",
            "声压级",
            "声级",
            "a-weighted",
            "a加权",
            "level",
            "db",
            "direct",
            "直达",
        ],
    )
    return {"x": x_col, "y": y_col, "z": z_col, "spl": spl_col}


def read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=sheet_name)
    col_map = detect_columns(df)
    out = df[[col_map["x"], col_map["y"], col_map["z"], col_map["spl"]]].copy()
    out.columns = ["x", "y", "z", "spl"]
    out = out.dropna(subset=["x", "y", "z", "spl"])
    out = out.astype({"x": float, "y": float, "z": float, "spl": float})
    return out.reset_index(drop=True)


def minmax_norm(arr: np.ndarray) -> np.ndarray:
    mn = arr.min(axis=0)
    mx = arr.max(axis=0)
    span = np.where((mx - mn) == 0, 1.0, (mx - mn))
    return (arr - mn) / span


def greedy_unique_match(src_xyz: np.ndarray, dst_xyz: np.ndarray) -> List[Tuple[int, int, float]]:
    n, m = src_xyz.shape[0], dst_xyz.shape[0]
    if n > m:
        raise ValueError(
            f"唯一匹配要求 source 点数 <= target 点数，当前 source={n}, target={m}。"
        )
    dmat = np.linalg.norm(src_xyz[:, None, :] - dst_xyz[None, :, :], axis=2)
    pairs = [(i, j, dmat[i, j]) for i in range(n) for j in range(m)]
    pairs.sort(key=lambda x: x[2])

    used_i = set()
    used_j = set()
    matched: List[Tuple[int, int, float]] = []
    for i, j, d in pairs:
        if i in used_i or j in used_j:
            continue
        matched.append((i, j, d))
        used_i.add(i)
        used_j.add(j)
        if len(matched) == n:
            break

    if len(matched) != n:
        raise RuntimeError("唯一匹配失败，未完成全量一一配对。")
    return matched


def nearest_repeat_match(src_xyz: np.ndarray, dst_xyz: np.ndarray) -> List[Tuple[int, int, float]]:
    dmat = np.linalg.norm(src_xyz[:, None, :] - dst_xyz[None, :, :], axis=2)
    nearest_j = np.argmin(dmat, axis=1)
    return [(int(i), int(j), float(dmat[i, j])) for i, j in enumerate(nearest_j)]


def assign_grid_labels(
    df: pd.DataFrame, row_count: int | None, col_count: int, y_tol: float
) -> pd.DataFrame:
    out = df.copy()
    out["__idx"] = np.arange(len(out))
    out = out.sort_values(["y", "x", "z"], ascending=[False, True, True]).reset_index(drop=True)
    max_full_rows = len(out) // col_count
    if max_full_rows <= 0:
        raise ValueError(f"点数不足：当前 {len(out)}，无法组成一整行 col_count={col_count}")

    effective_rows = min(row_count, max_full_rows) if row_count is not None else max_full_rows
    expected_total = effective_rows * col_count

    row_vals: List[float] = []
    row_ids: List[int] = []
    for yv in out["y"].to_numpy():
        matched_rid = None
        for rid, rv in enumerate(row_vals):
            if abs(float(yv) - rv) <= y_tol:
                matched_rid = rid
                break
        if matched_rid is None:
            row_vals.append(float(yv))
            row_ids.append(len(row_vals) - 1)
        else:
            row_ids.append(matched_rid)
    out["row_id"] = row_ids

    # 若按容差分层不稳定，则回退到“按 y 排序后每 col_count 个点切一行”。
    if len(row_vals) != effective_rows:
        out = out.sort_values(["y", "x", "z"], ascending=[False, True, True]).reset_index(drop=True)
        out = out.iloc[:expected_total].copy()
        out["row_id"] = out.index // col_count

    out["col_id"] = -1
    for rid in range(effective_rows):
        sub = out[out["row_id"] == rid].sort_values("x", ascending=True)
        if len(sub) < col_count:
            raise ValueError(
                f"第 {rid} 行点数={len(sub)}，小于期望 {col_count} 列；请确认数据是否完整。"
            )
        sub = sub.iloc[:col_count]
        out.loc[sub.index, "col_id"] = np.arange(col_count)

    out = out[out["col_id"] >= 0].copy()

    return out


def grid_index_match(
    df_self: pd.DataFrame,
    df_ease5: pd.DataFrame,
    row_count: int | None,
    col_count: int,
    y_tol: float,
) -> List[Tuple[int, int, float]]:
    if row_count is None:
        row_count = min(len(df_self) // col_count, len(df_ease5) // col_count)
        if row_count <= 0:
            raise ValueError("两侧数据都无法组成完整一行，请检查 col_count 设置。")

    s = assign_grid_labels(df_self, row_count=row_count, col_count=col_count, y_tol=y_tol)
    e = assign_grid_labels(df_ease5, row_count=row_count, col_count=col_count, y_tol=y_tol)

    s_map = {(int(r), int(c)): int(i) for r, c, i in zip(s["row_id"], s["col_id"], s["__idx"])}
    e_map = {(int(r), int(c)): int(i) for r, c, i in zip(e["row_id"], e["col_id"], e["__idx"])}

    keys = sorted(set(s_map.keys()) & set(e_map.keys()))
    if not keys:
        raise ValueError("grid 配对失败，没有可对齐的 (row_id, col_id)。")

    matched: List[Tuple[int, int, float]] = []
    for key in keys:
        i = s_map[key]
        j = e_map[key]
        ds = df_self.loc[i, ["x", "y", "z"]].to_numpy(dtype=float)
        de = df_ease5.loc[j, ["x", "y", "z"]].to_numpy(dtype=float)
        dist = float(np.linalg.norm(ds - de))
        matched.append((i, j, dist))
    return matched


def match_self_ease5(
    self_xyz: np.ndarray,
    ease5_xyz: np.ndarray,
    mode: str,
    anchor: str,
    df_self: pd.DataFrame | None = None,
    df_ease5: pd.DataFrame | None = None,
    row_count: int | None = None,
    col_count: int = 20,
    y_tol: float = 1e-3,
) -> List[Tuple[int, int, float]]:
    n_self, n_ease5 = self_xyz.shape[0], ease5_xyz.shape[0]

    if anchor not in {"self", "ease5", "smaller"}:
        raise ValueError(f"不支持的锚点策略: {anchor}")

    if anchor == "self":
        source_is_self = True
    elif anchor == "ease5":
        source_is_self = False
    else:
        source_is_self = n_self <= n_ease5

    src = self_xyz if source_is_self else ease5_xyz
    dst = ease5_xyz if source_is_self else self_xyz

    if mode == "grid":
        if df_self is None or df_ease5 is None:
            raise ValueError("grid 模式需要原始DataFrame。")
        return grid_index_match(
            df_self=df_self,
            df_ease5=df_ease5,
            row_count=row_count,
            col_count=col_count,
            y_tol=y_tol,
        )
    if mode == "repeat":
        matched_src_dst = nearest_repeat_match(src, dst)
    elif mode == "unique":
        if src.shape[0] <= dst.shape[0]:
            matched_src_dst = greedy_unique_match(src, dst)
        else:
            matched_src_dst = greedy_unique_match(dst, src)
            matched_src_dst = [(j, i, d) for i, j, d in matched_src_dst]
    else:
        raise ValueError(f"不支持的匹配模式: {mode}")

    if source_is_self:
        return matched_src_dst

    return [(idx_self, idx_ease5, dist) for idx_ease5, idx_self, dist in matched_src_dst]


def plot_outputs(
    merged: pd.DataFrame,
    out_dir: Path,
    title_prefix: str,
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    # 1) 声压级散点对比图
    plt.figure(figsize=(7, 7))
    plt.scatter(merged["spl_self"], merged["spl_ease5"], alpha=0.8, s=20)
    vmin = min(merged["spl_self"].min(), merged["spl_ease5"].min())
    vmax = max(merged["spl_self"].max(), merged["spl_ease5"].max())
    plt.plot([vmin, vmax], [vmin, vmax], "r--", linewidth=1.2, label="y=x")
    plt.xlabel("自研 SPL (dB)")
    plt.ylabel("EASE5 SPL (dB)")
    plt.title(f"{title_prefix} | SPL 对比散点图")
    plt.legend()
    plt.grid(alpha=0.2)
    plt.tight_layout()
    plt.savefig(out_dir / "01_spl_scatter.png", dpi=180)
    plt.close()

    # 2) 空间差值分布图（x-y 平面，颜色为差值）
    plt.figure(figsize=(8, 5))
    sc = plt.scatter(
        merged["x_self"],
        merged["y_self"],
        c=merged["delta_db"],
        cmap="coolwarm",
        s=35,
    )
    plt.colorbar(sc, label="SPL差值 (自研 - EASE5) dB")
    plt.xlabel("x (self)")
    plt.ylabel("y (self)")
    plt.title(f"{title_prefix} | 空间差值散点图")
    plt.grid(alpha=0.2)
    plt.tight_layout()
    plt.savefig(out_dir / "02_spatial_delta_scatter.png", dpi=180)
    plt.close()

    # 3) 差值直方图
    plt.figure(figsize=(8, 5))
    plt.hist(merged["delta_db"], bins=20, alpha=0.85, edgecolor="black")
    plt.xlabel("SPL差值 (自研 - EASE5) dB")
    plt.ylabel("数量")
    plt.title(f"{title_prefix} | 差值直方图")
    plt.grid(alpha=0.2)
    plt.tight_layout()
    plt.savefig(out_dir / "03_delta_hist.png", dpi=180)
    plt.close()


def write_excel_with_scatter_charts(merged: pd.DataFrame, stats: Dict[str, float], out_dir: Path) -> Path:
    out_path = out_dir / "matched_points_with_charts.xlsx"
    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="matched_points")
            pd.DataFrame([stats]).to_excel(writer, index=False, sheet_name="stats")
    except PermissionError:
        ts_name = f"matched_points_with_charts_{pd.Timestamp.now():%Y%m%d_%H%M%S}.xlsx"
        out_path = out_dir / ts_name
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            merged.to_excel(writer, index=False, sheet_name="matched_points")
            pd.DataFrame([stats]).to_excel(writer, index=False, sheet_name="stats")

    wb = load_workbook(out_path)
    ws = wb["matched_points"]

    # 冻结首行并开启表头筛选
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    headers = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}
    max_row = ws.max_row

    # 图1：SPL 对比散点图（self vs ease5）
    chart1 = ScatterChart()
    chart1.title = "SPL 对比散点图"
    chart1.x_axis.title = "自研 SPL (dB)"
    chart1.y_axis.title = "EASE5 SPL (dB)"
    chart1.style = 2
    xvalues_1 = Reference(ws, min_col=headers["spl_self"], min_row=2, max_row=max_row)
    yvalues_1 = Reference(ws, min_col=headers["spl_ease5"], min_row=2, max_row=max_row)
    series1 = Series(yvalues_1, xvalues_1, title="点位")
    chart1.series.append(series1)
    ws.add_chart(chart1, "N2")

    # 图2：空间差值散点图（x_self vs y_self，先做普通散点，颜色映射保留在 png）
    chart2 = ScatterChart()
    chart2.title = "空间散点图（x-y）"
    chart2.x_axis.title = "x (self)"
    chart2.y_axis.title = "y (self)"
    chart2.style = 2
    xvalues_2 = Reference(ws, min_col=headers["x_self"], min_row=2, max_row=max_row)
    yvalues_2 = Reference(ws, min_col=headers["y_self"], min_row=2, max_row=max_row)
    series2 = Series(yvalues_2, xvalues_2, title="点位坐标")
    chart2.series.append(series2)
    ws.add_chart(chart2, "N20")

    wb.save(out_path)
    return out_path


def build_row_col_layout(df: pd.DataFrame, col_count: int) -> pd.DataFrame:
    out = df.copy()
    out["src_idx"] = np.arange(len(out))
    out = out.sort_values(["y", "x", "z"], ascending=[False, True, True]).reset_index(drop=True)

    row_ids: List[int] = []
    col_ids: List[int] = []
    seq_ids: List[int] = []
    seq = 1
    for i in range(len(out)):
        row_id = (i // col_count) + 1
        col_id = (i % col_count) + 1
        row_ids.append(row_id)
        col_ids.append(col_id)
        seq_ids.append(seq)
        seq += 1

    out["序号"] = seq_ids
    out["行号"] = row_ids
    out["列号"] = col_ids
    return out


def match_by_ease5_row_nearest_column(
    df_self: pd.DataFrame,
    df_ease5: pd.DataFrame,
    col_count: int,
) -> pd.DataFrame:
    from scipy.optimize import linear_sum_assignment

    # 直接使用原始数据中的行号、列号，不再重新排序
    # 提取关键列（通过列索引，避免编码问题）
    def _extract(df):
        return pd.DataFrame({
            "序号": df.iloc[:, 0].astype(int),
            "行号": df.iloc[:, 1].astype(int),
            "列号": df.iloc[:, 2].astype(int),
            "x": df.iloc[:, 5].astype(float),
            "y": df.iloc[:, 6].astype(float),
            "z": df.iloc[:, 7].astype(float),
            "spl": df.iloc[:, 8].astype(float),
        })

    s = _extract(df_self)
    e = _extract(df_ease5)

    # 过滤无效点（x=0, y=0, z=0 的是占位数据）
    s = s[~((s["x"] == 0) & (s["y"] == 0) & (s["z"] == 0))].copy()
    e = e[~((e["x"] == 0) & (e["y"] == 0) & (e["z"] == 0))].copy()

    common_rows = sorted(set(s["行号"]) & set(e["行号"]))
    if not common_rows:
        raise ValueError("两侧数据没有公共行号，无法匹配。")

    rows = []
    for rid in common_rows:
        s_row = s[s["行号"] == rid].reset_index(drop=True)
        e_row = e[e["行号"] == rid].reset_index(drop=True)

        # 计算该行内所有 self-ease5 点对的欧氏距离矩阵
        s_xyz = s_row[["x", "y", "z"]].to_numpy(dtype=float)
        e_xyz = e_row[["x", "y", "z"]].to_numpy(dtype=float)
        dist_mat = np.linalg.norm(s_xyz[:, None, :] - e_xyz[None, :, :], axis=2)

        # 匈牙利算法求最优一一匹配
        row_idx, col_idx = linear_sum_assignment(dist_mat)

        for si, ei in zip(row_idx, col_idx):
            sp = s_row.iloc[si]
            ep = e_row.iloc[ei]
            dist = float(dist_mat[si, ei])
            rows.append({
                "ease5_序号": int(ep["序号"]),
                "ease5_行号": int(ep["行号"]),
                "ease5_列号": int(ep["列号"]),
                "ease5_x": float(ep["x"]),
                "ease5_y": float(ep["y"]),
                "ease5_z": float(ep["z"]),
                "self_序号": int(sp["序号"]),
                "self_行号": int(sp["行号"]),
                "self_列号": int(sp["列号"]),
                "self_x": float(sp["x"]),
                "self_y": float(sp["y"]),
                "self_z": float(sp["z"]),
                "ease5_spl": float(ep["spl"]),
                "self_spl": float(sp["spl"]),
                "delta_db(self-ease5)": float(sp["spl"] - ep["spl"]),
                "match_dist_xyz": dist,
            })

    return pd.DataFrame(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="对比两个Excel指定sheet中的 xyz 点位声压级（坐标允许略有偏差）。"
    )
    parser.add_argument("--self_file", required=False, help="自研软件导出的xlsx路径")
    parser.add_argument("--ease5_file", required=False, help="EASE5导出的xlsx路径")
    parser.add_argument("--self_sheet", default=None, help="自研Excel的sheet名")
    parser.add_argument("--ease5_sheet", default=None, help="EASE5 Excel的sheet名")
    parser.add_argument("--out_dir", default=None, help="输出目录")
    parser.add_argument(
        "--match_mode",
        default=None,
        choices=["grid", "repeat", "unique"],
        help="点位配对模式：grid=按行列索引对齐；repeat=最近邻可重复；unique=最近邻唯一配对",
    )
    parser.add_argument(
        "--anchor",
        default=None,
        choices=["self", "ease5", "smaller"],
        help="以哪一侧作为匹配源：smaller=自动选择点数更少的一侧",
    )
    parser.add_argument("--row_count", type=int, default=None, help="grid模式行数，留空自动")
    parser.add_argument("--col_count", type=int, default=None, help="grid模式列数")
    parser.add_argument("--y_tol", type=float, default=None, help="grid模式按y分层容差")
    parser.add_argument("--generate_charts", action="store_true", help="是否生成图表（默认不生成）")
    args = parser.parse_args()

    def auto_pick_xlsx(role: str) -> Path:
        cwd = Path.cwd()
        candidates = sorted(cwd.rglob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if not candidates:
            raise FileNotFoundError("当前目录及子目录未找到任何 xlsx 文件。")

        if role == "self":
            keywords = ["声压数据", "self", "自研"]
        else:
            keywords = ["es5数据", "ease5", "es5"]

        for p in candidates:
            name = p.name.lower()
            if all(k.lower() in name for k in [keywords[0]]):
                return p
        for p in candidates:
            name = p.name.lower()
            if any(k.lower() in name for k in keywords):
                return p

        raise FileNotFoundError(f"未自动识别到 {role} 文件，请手动通过参数指定。")

    self_file = args.self_file or SELF_FILE_CONFIG
    ease5_file = args.ease5_file or EASE5_FILE_CONFIG
    self_sheet = args.self_sheet or SELF_SHEET_CONFIG
    ease5_sheet = args.ease5_sheet or EASE5_SHEET_CONFIG
    out_dir = Path(args.out_dir or OUT_DIR_CONFIG)
    match_mode = args.match_mode or MATCH_MODE_CONFIG
    anchor = args.anchor or ANCHOR_CONFIG
    row_count = args.row_count if args.row_count is not None else ROW_COUNT_CONFIG
    col_count = args.col_count if args.col_count is not None else COL_COUNT_CONFIG
    y_tol = args.y_tol if args.y_tol is not None else Y_TOL_CONFIG
    generate_charts = args.generate_charts or GENERATE_CHARTS_CONFIG

    self_path = Path(self_file) if str(self_file).strip() else auto_pick_xlsx("self")
    ease5_path = Path(ease5_file) if str(ease5_file).strip() else auto_pick_xlsx("ease5")

    # 读取原始数据（保留行号、列号等完整信息）
    df_self_raw = pd.read_excel(self_path, sheet_name=self_sheet)
    df_ease5_raw = pd.read_excel(ease5_path, sheet_name=ease5_sheet)

    merged = match_by_ease5_row_nearest_column(
        df_self=df_self_raw,
        df_ease5=df_ease5_raw,
        col_count=col_count,
    )
    out_dir.mkdir(parents=True, exist_ok=True)
    merged.to_csv(out_dir / "matched_points_delta.csv", index=False, encoding="utf-8-sig")

    stats = {
        "count": int(len(merged)),
        "mean_delta_db": float(merged["delta_db(self-ease5)"].mean()),
        "mae_db": float(np.abs(merged["delta_db(self-ease5)"]).mean()),
        "rmse_db": float(np.sqrt(np.mean(merged["delta_db(self-ease5)"] ** 2))),
        "p95_abs_db": float(np.percentile(np.abs(merged["delta_db(self-ease5)"]), 95)),
        "max_abs_db": float(np.abs(merged["delta_db(self-ease5)"]).max()),
        "mean_match_dist_xyz": float(merged["match_dist_xyz"].mean()),
        "max_match_dist_xyz": float(merged["match_dist_xyz"].max()),
    }
    pd.DataFrame([stats]).to_csv(out_dir / "stats_summary.csv", index=False, encoding="utf-8-sig")

    excel_out = out_dir / "matched_points_with_charts.xlsx"
    with pd.ExcelWriter(excel_out, engine="openpyxl") as writer:
        merged.to_excel(writer, index=False, sheet_name="matched_points")
        pd.DataFrame([stats]).to_excel(writer, index=False, sheet_name="stats")

    if generate_charts:
        plot_outputs(
            merged=pd.DataFrame(
                {
                    "spl_self": merged["self_spl"],
                    "spl_ease5": merged["ease5_spl"],
                    "x_self": merged["self_x"],
                    "y_self": merged["self_y"],
                    "delta_db": merged["delta_db(self-ease5)"],
                }
            ),
            out_dir=out_dir,
            title_prefix=f"{self_path.stem} vs {ease5_path.stem}",
        )
        excel_out = write_excel_with_scatter_charts(
            merged=pd.DataFrame(
                {
                    "spl_self": merged["self_spl"],
                    "spl_ease5": merged["ease5_spl"],
                    "x_self": merged["self_x"],
                    "y_self": merged["self_y"],
                    "delta_db": merged["delta_db(self-ease5)"],
                }
            ),
            stats=stats,
            out_dir=out_dir,
        )

    print("对比完成。输出文件：")
    print(f"- {out_dir / 'matched_points_delta.csv'}")
    print(f"- {out_dir / 'stats_summary.csv'}")
    print(f"- {excel_out}")
    if generate_charts:
        print(f"- {out_dir / '01_spl_scatter.png'}")
        print(f"- {out_dir / '02_spatial_delta_scatter.png'}")
        print(f"- {out_dir / '03_delta_hist.png'}")


if __name__ == "__main__":
    main()
