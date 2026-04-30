"""
ES5数据采集工具
功能：
1. 手动选取目标软件的矩形区域
2. 在区域内生成20x10的网格点
3. 控制鼠标依次移动到各点（首次移动到第一个点时执行Shift+滚轮向上）
4. 抓取坐标和声压级信息
5. 写入Excel文件
"""


import cv2
import numpy as np
import pyautogui
import time
from openpyxl import Workbook
from datetime import datetime
import ctypes
import sys

# 配置tesseract路径
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r'E:\Program Files\Tesseract-OCR\tesseract.exe'

# 禁用pyautogui的安全检查
pyautogui.PAUSE = 0.1
pyautogui.FAILSAFE = False


class ES5DataCollector:
    def __init__(self):
        self.selection_rect = None
        self.status_bar_rect = None  # 顶部状态栏区域
        self.grid_points = []
        self.collected_data = []
        self.cols = 20  # 列数
        self.rows = 10  # 行数
        self.first_point = True  # 标记是否是第一个点

    def select_region(self):
        """
        让用户在屏幕上选择一个矩形区域
        返回: (x1, y1, x2, y2) - 矩形的左上角和右下角坐标
        """
        print("\n" + "="*60)
        print("区域选择模式")
        print("="*60)
        print("请按以下步骤操作：")
        print("1. 确保目标软件窗口已打开并可见")
        print("2. 在接下来的3秒内，切换到目标软件窗口")
        print("3. 使用鼠标左键拖拽选择要采集数据的矩形区域")
        print("4. 选择完成后，按ESC键或右键结束选择")
        print("="*60 + "\n")

        # 等待用户切换窗口
        print("3秒后开始截图...")
        time.sleep(3)

        # 获取屏幕截图
        screenshot = pyautogui.screenshot()
        screenshot_np = np.array(screenshot)
        screenshot_rgb = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2BGR)

        # 创建窗口
        window_name = "选择数据采集区域 (拖拽选择, ESC/右键结束)"
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
        cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

        # 初始化变量
        self.selection_rect = None
        drawing = False
        start_x, start_y = 0, 0

        def mouse_callback(event, x, y, flags, param):
            nonlocal drawing, start_x, start_y

            if event == cv2.EVENT_LBUTTONDOWN:
                drawing = True
                start_x, start_y = x, y
                self.selection_rect = None

            elif event == cv2.EVENT_MOUSEMOVE:
                if drawing:
                    # 显示选择框
                    display = screenshot_rgb.copy()
                    cv2.rectangle(display, (start_x, start_y), (x, y), (0, 255, 0), 2)
                    cv2.imshow(window_name, display)

            elif event == cv2.EVENT_LBUTTONUP:
                drawing = False
                end_x, end_y = x, y
                # 确保坐标顺序正确
                x1, x2 = min(start_x, end_x), max(start_x, end_x)
                y1, y2 = min(start_y, end_y), max(start_y, end_y)
                self.selection_rect = (x1, y1, x2, y2)

            elif event == cv2.EVENT_RBUTTONUP:
                cv2.destroyAllWindows()

        cv2.setMouseCallback(window_name, mouse_callback)
        cv2.imshow(window_name, screenshot_rgb)

        print("请在屏幕上拖拽选择区域...")

        while True:
            key = cv2.waitKey(1) & 0xFF
            if key == 27:  # ESC键
                break
            if self.selection_rect is not None and not drawing:
                break

        cv2.destroyAllWindows()

        if self.selection_rect:
            x1, y1, x2, y2 = self.selection_rect
            width = x2 - x1
            height = y2 - y1
            print(f"\n已选择区域: ({x1}, {y1}) 到 ({x2}, {y2})")
            print(f"区域大小: {width} x {height} 像素")
            return self.selection_rect
        else:
            print("\n未选择有效区域！")
            return None

    def select_status_bar_region(self):
        """
        让用户选择顶部状态栏区域（用于OCR识别坐标和声压级）
        返回: (x1, y1, x2, y2) - 矩形的左上角和右下角坐标
        """
        print("\n" + "="*60)
        print("状态栏区域选择")
        print("="*60)
        print("请按以下步骤操作：")
        print("1. 在接下来的3秒内，切换到目标软件窗口")
        print("2. 使用鼠标左键拖拽选择顶部状态栏区域")
        print("   (应包含坐标和声压级信息)")
        print("3. 选择完成后，按ESC键或右键结束选择")
        print("="*60 + "\n")

        # 等待用户切换窗口
        print("3秒后开始截图...")
        time.sleep(3)

        # 获取屏幕截图
        screenshot = pyautogui.screenshot()
        screenshot_np = np.array(screenshot)
        screenshot_rgb = cv2.cvtColor(screenshot_np, cv2.COLOR_RGB2BGR)

        # 创建窗口
        window_name = "选择状态栏区域 (拖拽选择, ESC/右键结束)"
        cv2.namedWindow(window_name, cv2.WINDOW_NORMAL)
        cv2.setWindowProperty(window_name, cv2.WND_PROP_FULLSCREEN, cv2.WINDOW_FULLSCREEN)

        # 初始化变量
        self.status_bar_rect = None
        drawing = False
        start_x, start_y = 0, 0

        def mouse_callback(event, x, y, flags, param):
            nonlocal drawing, start_x, start_y

            if event == cv2.EVENT_LBUTTONDOWN:
                drawing = True
                start_x, start_y = x, y
                self.status_bar_rect = None

            elif event == cv2.EVENT_MOUSEMOVE:
                if drawing:
                    # 显示选择框
                    display = screenshot_rgb.copy()
                    cv2.rectangle(display, (start_x, start_y), (x, y), (255, 0, 0), 2)
                    cv2.imshow(window_name, display)

            elif event == cv2.EVENT_LBUTTONUP:
                drawing = False
                end_x, end_y = x, y
                # 确保坐标顺序正确
                x1, x2 = min(start_x, end_x), max(start_x, end_x)
                y1, y2 = min(start_y, end_y), max(start_y, end_y)
                self.status_bar_rect = (x1, y1, x2, y2)

            elif event == cv2.EVENT_RBUTTONUP:
                cv2.destroyAllWindows()

        cv2.setMouseCallback(window_name, mouse_callback)
        cv2.imshow(window_name, screenshot_rgb)

        print("请在屏幕上拖拽选择状态栏区域...")

        while True:
            key = cv2.waitKey(1) & 0xFF
            if key == 27:  # ESC键
                break
            if self.status_bar_rect is not None and not drawing:
                break

        cv2.destroyAllWindows()

        if self.status_bar_rect:
            x1, y1, x2, y2 = self.status_bar_rect
            width = x2 - x1
            height = y2 - y1
            print(f"\n已选择状态栏区域: ({x1}, {y1}) 到 ({x2}, {y2})")
            print(f"区域大小: {width} x {height} 像素")
            return self.status_bar_rect
        else:
            print("\n未选择有效区域！")
            return None

    def generate_grid_points(self, x1, y1, x2, y2):
        """
        在指定矩形区域内生成均匀分布的网格点
        """
        width = x2 - x1
        height = y2 - y1

        # 计算网格间距
        x_step = width / self.cols
        y_step = height / self.rows

        points = []
        for row in range(self.rows):
            for col in range(self.cols):
                # 计算每个点的中心位置
                x = int(x1 + col * x_step + x_step / 2)
                y = int(y1 + row * y_step + y_step / 2)
                points.append((x, y, row + 1, col + 1))

        self.grid_points = points
        print(f"\n已生成 {len(points)} 个网格点 ({self.rows}行 x {self.cols}列)")
        return points

    def _check_dB_exists(self):
        """
        检查状态栏区域是否存在dB字符
        返回: True表示存在dB，False表示不存在
        """
        try:
            from PIL import Image, ImageEnhance

            # 截取整个屏幕
            screenshot = pyautogui.screenshot()

            # 使用用户选择的状态栏区域进行截图
            if self.status_bar_rect:
                x1, y1, x2, y2 = self.status_bar_rect
                top_region = screenshot.crop((x1, y1, x2, y2))
            else:
                screen_width, _ = pyautogui.size()
                top_region = screenshot.crop((0, 0, screen_width, 100))

            # 图像预处理
            top_region = top_region.convert('L')
            top_region = ImageEnhance.Contrast(top_region).enhance(3.0)
            top_region = ImageEnhance.Sharpness(top_region).enhance(3.0)

            # OCR识别
            text = pytesseract.image_to_string(top_region, config='--psm 6 --oem 3')
            print(f"检查dB - OCR文本: {text.strip()}")

            # 检查是否包含dB
            return 'dB' in text or 'd' in text
        except Exception as e:
            print(f"检查dB失败: {e}")
            return False

    def get_screen_info_at_position(self, x, y):
        """
        获取指定屏幕位置的坐标和声压级信息
        通过OCR从状态栏区域读取
        """
        # 移动鼠标到指定位置
        pyautogui.moveTo(x, y)
        time.sleep(0.5)  # 等待软件更新显示

        # 如果是第一个点，执行Shift+滚轮操作，直到出现dB
        if self.first_point:
            self._wait_for_dB_appear()
            self.first_point = False

        # 使用OCR从状态栏区域读取坐标和声压级信息
        # 如果没有识别到有效的dB值，重试Shift+滚轮操作
        info = self._get_info_with_retry()
        return info

    def _wait_for_dB_appear(self):
        """
        首次移动到采集点时，执行Shift+滚轮操作直到出现dB字符
        """
        print("首次移动到采集点，执行Shift+滚轮操作...")
        max_attempts = 10  # 最大尝试次数

        for attempt in range(max_attempts):
            # 尝试Shift+滚轮向上
            self._scroll_with_shift('up')

            # 检查是否出现dB
            if self._check_dB_exists():
                print(f"检测到dB字符，执行成功！")
                return True

            # Shift+滚轮向上无效，尝试Shift+滚轮向下
            print(f"Shift+滚轮向上无效，尝试Shift+滚轮向下...")
            self._scroll_with_shift('down')

            if self._check_dB_exists():
                print(f"检测到dB字符，执行成功！")
                return True

            if attempt < max_attempts - 1:
                print(f"未检测到dB字符，第{attempt + 1}次重试...")

        print("警告: 多次尝试后仍未检测到dB字符，继续执行...")
        return False

    def _get_info_with_retry(self):
        """
        获取OCR信息，如果识别失败则重试Shift+滚轮操作
        """
        max_retry = 3

        for retry in range(max_retry):
            info = self._get_info_by_ocr()

            # 检查是否识别到了有效的声压级（dB值在合理范围内）
            if info['spl'] > 0:
                return info

            if retry < max_retry - 1:
                print(f"未识别到有效dB值，第{retry + 1}次重试，执行Shift+滚轮向上...")

                # 先尝试Shift+滚轮向上
                self._scroll_with_shift('up')
                info = self._get_info_by_ocr()
                if info['spl'] > 0:
                    return info

                print(f"Shift+滚轮向上无效，执行Shift+滚轮向下...")
                # 再尝试Shift+滚轮向下
                self._scroll_with_shift('down')
                info = self._get_info_by_ocr()
                if info['spl'] > 0:
                    return info

        print("警告: 多次尝试后仍未识别到有效dB值")
        return info

    def _scroll_with_shift(self, direction='up'):
        """
        执行Shift+滚轮操作
        direction: 'up' 向上滚动, 'down' 向下滚动
        """
        pyautogui.keyDown('shift')
        time.sleep(0.1)
        if direction == 'up':
            pyautogui.scroll(10)
        else:
            pyautogui.scroll(-10)
        time.sleep(0.1)
        pyautogui.keyUp('shift')
        time.sleep(0.5)

    def _get_info_by_ocr(self):
        """
        使用OCR从屏幕读取坐标和声压级信息
        使用用户选择的状态栏区域进行截图
        """
        try:
            import re
            from PIL import Image, ImageEnhance, ImageFilter, ImageOps

            # 截取整个屏幕
            screenshot = pyautogui.screenshot()

            # 使用用户选择的状态栏区域进行截图
            if self.status_bar_rect:
                x1, y1, x2, y2 = self.status_bar_rect
                top_region = screenshot.crop((x1, y1, x2, y2))
            else:
                # 如果没有选择状态栏区域，使用默认顶部区域
                screen_width, _ = pyautogui.size()
                top_region = screenshot.crop((0, 0, screen_width, 100))

            # 图像预处理：增强负号等细字符
            top_region = top_region.convert('L')  # 转为灰度图
            top_region = ImageOps.autocontrast(top_region)
            top_region = top_region.resize((top_region.width * 3, top_region.height * 3), Image.Resampling.LANCZOS)
            top_region = ImageEnhance.Contrast(top_region).enhance(4.0)
            top_region = ImageEnhance.Sharpness(top_region).enhance(3.5)
            top_region = top_region.point(lambda p: 255 if p > 170 else 0)

            # 尝试多种OCR配置，选择最佳结果
            configs = [
                # LSTM引擎，多行文本
                '--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789.-+|XYdB: ',
                # LSTM引擎，单行文本
                '--psm 7 --oem 3 -c tessedit_char_whitelist=0123456789.-+|XYdB: ',
                # 神经网络引擎，多行文本
                '--psm 6 --oem 1 -c tessedit_char_whitelist=0123456789.-+|XYdB: ',
                # 神经网络引擎，单行文本
                '--psm 7 --oem 1 -c tessedit_char_whitelist=0123456789.-+|XYdB: ',
                # 传统的引擎
                '--psm 6 --oem 0 -c tessedit_char_whitelist=0123456789.-+|XYdB: '
            ]

            best_text = ""
            best_score = 0

            for config in configs:
                try:
                    text = pytesseract.image_to_string(top_region, config=config)
                    # 计算识别质量分数（优先保留负号和小数点）
                    decimal_count = text.count('.')
                    minus_count = text.count('-')
                    digit_count = len(re.findall(r'\d', text))
                    score = minus_count * 20 + decimal_count * 10 + digit_count
                    if score > best_score:
                        best_score = score
                        best_text = text
                except:
                    continue

            text = best_text if best_text else pytesseract.image_to_string(top_region, config='--psm 6 --oem 3 -c tessedit_char_whitelist=0123456789.-+|XYdB: ')
            print(f"OCR识别文本: {text.strip()}")

            # 解析格式：25.67 | 7.72 | 1.27 (三个坐标用|分隔)
            # 可能有三行：坐标行、Area行、声压级行
            # 需要过滤掉包含Area的行
            lines = text.strip().split('\n')

            # 过滤掉包含Area的行
            filtered_lines = [line for line in lines if 'Area' not in line and 'area' not in line]

            x_coord = 0.0
            y_coord = 0.0
            z_coord = 0.0
            spl = 0.0

            # 解析第一行（过滤后）的坐标
            if filtered_lines:
                # 尝试用竖线分隔
                coords = [c.strip() for c in filtered_lines[0].split('|')]
                # 如果没有竖线，尝试按空格分隔
                if len(coords) == 1:
                    coords = [c.strip() for c in filtered_lines[0].split() if c.strip()]

                # 提取数字
                numbers = []
                for coord in coords:
                    # 提取数字（包括负数和小数）
                    nums = re.findall(r'[-+]?\d*\.\d+|[-+]?\d+', coord)
                    numbers.extend(nums)

                # 后处理：智能修复缺失的小数点
                if len(numbers) >= 1:
                    x_coord = self._fix_decimal_point(numbers[0])
                if len(numbers) >= 2:
                    y_coord = self._fix_decimal_point(numbers[1])
                if len(numbers) >= 3:
                    z_coord = self._fix_decimal_point(numbers[2])

            # 解析声压级：从过滤后的行中找包含dB或数字+单位格式的行
            for line in filtered_lines[1:]:
                line = line.strip()
                # 匹配声压级格式：数字+dB或数字+d（支持负数）
                spl_match = re.search(r'([-+]?\d+\.?\d*)\s*dB', line, re.IGNORECASE)
                if spl_match:
                    spl = float(spl_match.group(1))
                    break
                # 如果没有dB，尝试匹配较大的数值（声压级通常在-100到100之间）
                nums = re.findall(r'[-+]?\d+\.\d+|[-+]?\d+', line)
                for num_str in nums:
                    try:
                        num_val = float(num_str)
                        if -100 <= num_val <= 100:  # 合理的声压级范围
                            spl = num_val
                            break
                    except:
                        pass

            return {
                'x_coord': x_coord,
                'coord_y': y_coord,
                'z_coord': z_coord,
                'spl': spl
            }
        except Exception as e:
            print(f"OCR识别错误: {e}")
            return {'x_coord': 0.0, 'coord_y': 0.0, 'z_coord': 0.0, 'spl': 0.0}

    def _fix_decimal_point(self, num_str):
        """
        智能修复缺失的小数点
        例如：750 -> 7.50, 4175 -> 41.75
        """
        try:
            # 如果已经有小数点，直接返回
            if '.' in num_str:
                return float(num_str)

            sign = -1 if num_str.startswith('-') else 1
            unsigned_num_str = num_str.lstrip('+-')

            # 如果数字太长，可能缺少小数点
            num = int(unsigned_num_str)
            if num > 100:
                # 尝试在倒数第二位插入小数点
                if len(unsigned_num_str) >= 3:
                    # 尝试不同的小数点位置
                    for pos in range(len(unsigned_num_str) - 2, 0, -1):
                        fixed = unsigned_num_str[:pos] + '.' + unsigned_num_str[pos:]
                        value = sign * float(fixed)
                        # 检查值是否在合理范围内（假设坐标和声压级在-100到100之间）
                        if -100 <= value <= 100:
                            return value
            return float(num_str)
        except:
            return 0.0

    def collect_data(self):
        """
        开始数据采集
        """
        if not self.grid_points:
            print("错误: 请先选择区域并生成网格点！")
            return False

        print("\n" + "="*60)
        print("开始数据采集")
        print("="*60)
        print(f"将采集 {len(self.grid_points)} 个点的数据")
        print("提示: 采集过程中请勿移动鼠标或切换窗口")
        print("注意: 首次移动到第一个采集点时会执行Shift+滚轮向上操作")
        print("="*60 + "\n")

        # 等待用户准备
        input("请确保目标软件窗口已激活，然后按回车键开始采集...")

        total = len(self.grid_points)
        self.collected_data = []

        for i, (screen_x, screen_y, row, col) in enumerate(self.grid_points):
            # 获取坐标和声压级信息（内部会移动鼠标），带重试
            info = None
            for attempt in range(3):
                info = self.get_screen_info_at_position(screen_x, screen_y)
                # 验证：坐标和声压级不能全为0，且不能有异常大的值
                if (info['x_coord'] != 0.0 or info['coord_y'] != 0.0 or
                    info['z_coord'] != 0.0 or info['spl'] != 0.0):
                    break
                if attempt < 2:
                    print(f"  点{i+1}识别结果全为0，重试({attempt+1}/3)...")
                    time.sleep(0.3)

            # 保存数据
            data_point = {
                'index': i + 1,
                'row': row,
                'col': col,
                'screen_x': screen_x,
                'screen_y': screen_y,
                'coord_x': info['x_coord'],
                'coord_y': info['coord_y'],
                'z_coord': info.get('z_coord', 0.0),
                'spl': info['spl']
            }
            self.collected_data.append(data_point)

            # 显示每个点的采集结果
            print(f"  点{i+1}: X={data_point['coord_x']}, Y={data_point['coord_y']}, Z={data_point['z_coord']}, SPL={data_point['spl']}")

            # 显示进度
            if (i + 1) % 10 == 0 or (i + 1) == total:
                print(f"进度: {i + 1}/{total} ({(i + 1) * 100 // total}%)")

        print(f"\n数据采集完成！共采集 {len(self.collected_data)} 个数据点")

        # 验证坐标一致性
        self.validate_coordinates()

        return True

    def validate_coordinates(self):
        """
        验证采集数据的坐标一致性
        - 检查每行的Y轴坐标是否一致
        - 检查每列的X轴坐标是否一致
        """
        if not self.collected_data:
            print("警告: 没有数据可验证！")
            return False

        print("\n" + "="*60)
        print("坐标一致性验证")
        print("="*60)

        has_error = False

        # 按行检查Y轴坐标
        print("\n[1] 按行检查Y轴坐标一致性:")
        print("-" * 60)

        for row_num in range(1, self.rows + 1):
            # 获取该行的所有数据点
            row_points = [d for d in self.collected_data if d['row'] == row_num]

            if not row_points:
                continue

            # 获取第一个点的Y坐标作为参考
            ref_y = row_points[0]['coord_y']
            inconsistent_points = []

            # 检查该行所有点的Y坐标
            for point in row_points:
                if abs(point['coord_y'] - ref_y) > 0.01:  # 允许0.01的误差
                    inconsistent_points.append(point)
                    has_error = True

            if inconsistent_points:
                print(f"\n  ⚠ 行 {row_num} 的Y轴坐标不一致 (参考值: {ref_y:.2f}):")
                for pt in inconsistent_points:
                    print(f"    - 序号:{pt['index']}, 行:{pt['row']}, 列:{pt['col']}, "
                          f"Y={pt['coord_y']:.2f} (偏差: {abs(pt['coord_y'] - ref_y):.2f})")
            else:
                print(f"  ✓ 行 {row_num}: Y={ref_y:.2f} (一致)")

        # 按列检查X轴坐标
        print("\n[2] 按列检查X轴坐标一致性:")
        print("-" * 60)

        for col_num in range(1, self.cols + 1):
            # 获取该列的所有数据点
            col_points = [d for d in self.collected_data if d['col'] == col_num]

            if not col_points:
                continue

            # 获取第一个点的X坐标作为参考
            ref_x = col_points[0]['coord_x']
            inconsistent_points = []

            # 检查该列所有点的X坐标
            for point in col_points:
                if abs(point['coord_x'] - ref_x) > 0.01:  # 允许0.01的误差
                    inconsistent_points.append(point)
                    has_error = True

            if inconsistent_points:
                print(f"\n  ⚠ 列 {col_num} 的X轴坐标不一致 (参考值: {ref_x:.2f}):")
                for pt in inconsistent_points:
                    print(f"    - 序号:{pt['index']}, 行:{pt['row']}, 列:{pt['col']}, "
                          f"X={pt['coord_x']:.2f} (偏差: {abs(pt['coord_x'] - ref_x):.2f})")
            else:
                print(f"  ✓ 列 {col_num}: X={ref_x:.2f} (一致)")

        # 总结
        print("\n" + "="*60)
        if has_error:
            print("验证结果: ⚠ 发现坐标不一致问题，请检查上方详细信息")
        else:
            print("验证结果: ✓ 所有坐标一致性检查通过")
        print("="*60 + "\n")

        return not has_error

    def save_to_excel(self, filename=None):
        """
        将采集的数据保存到Excel文件
        """
        if not self.collected_data:
            print("错误: 没有数据可保存！")
            return False

        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ES5数据_{timestamp}.xlsx"

        print(f"\n正在保存数据到 {filename}...")

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "ES5数据"

        # 写入表头
        headers = ['序号', '行号', '列号', '屏幕X', '屏幕Y', '坐标X', '坐标Y', '坐标Z', '声压级(dB)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, data in enumerate(self.collected_data, 2):
            ws.cell(row=row_idx, column=1, value=data['index'])
            ws.cell(row=row_idx, column=2, value=data['row'])
            ws.cell(row=row_idx, column=3, value=data['col'])
            ws.cell(row=row_idx, column=4, value=data['screen_x'])
            ws.cell(row=row_idx, column=5, value=data['screen_y'])
            ws.cell(row=row_idx, column=6, value=data['coord_x'])
            ws.cell(row=row_idx, column=7, value=data['coord_y'])
            ws.cell(row=row_idx, column=8, value=data.get('z_coord', 0.0))
            ws.cell(row=row_idx, column=9, value=data['spl'])

        # 保存文件
        wb.save(filename)
        print(f"数据已成功保存到: {filename}")
        return True

    def run(self):
        """
        运行完整的数据采集流程
        """
        print("\n" + "="*60)
        print("ES5数据采集工具")
        print("="*60)

        # 步骤1: 选择热力图区域
        rect = self.select_region()
        if not rect:
            print("热力图区域选择失败，程序退出")
            return

        # 步骤2: 选择状态栏区域（用于OCR识别）
        status_bar_rect = self.select_status_bar_region()
        if not status_bar_rect:
            print("状态栏区域选择失败，程序退出")
            return

        # 步骤3: 生成网格点
        x1, y1, x2, y2 = rect
        self.generate_grid_points(x1, y1, x2, y2)

        # 步骤4: 采集数据（自动移动鼠标并读取左上角数据）
        if self.collect_data():
            # 步骤5: 保存数据
            self.save_to_excel()

        print("\n程序执行完成！")


def main():
    collector = ES5DataCollector()
    collector.run()


if __name__ == "__main__":
    main()
